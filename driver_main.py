# -*- coding: utf-8 -*-
"""
Created on Wed Nov 27 11:59:15 2024

@author: Shephali.Gupta
"""

##---------------------------------------------------------------------------
# Load All Required Library
import numpy as np
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
import pandas as pd
import re
import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border
from concurrent.futures import ThreadPoolExecutor, as_completed
import os
from itertools import combinations
import multiprocessing as mp
import math 

# Load a pre-trained SentenceTransformer model
model = SentenceTransformer('all-MiniLM-L6-v2')
 
TEST_STEP_WEIGHT = 0.80
TEST_CASE_WEIGHT = 0.20
 
##---------------------------------------------------------------------------
 
def remove_quotes(text):
    text = text.lower().strip()
    text = re.sub(r'(step\s*\d+\s*[-:|s]*)', '', text, flags=re.IGNORECASE)
    #text = re.sub(r'Step \d+[-:.]', '', text)
    #text = re.sub(r'[sS]tep\s*\d+\s*[-:.|s]*', '', text)
    text = re.sub(r'^\s*\d+[-|s]*.\s*', '', text)
    text = text.strip()
    text = text.replace('“', '"').replace('”', '"')  # Standardize quotes
    text = text.replace("‘", "'").replace("’", "'")  # Replace single quotes with double quotes
    text = text.replace('\'', '"')
    text = re.sub(r'\s+', ' ', text)
    text = text.lower().strip()
    return text
 
def preprocess(texts):
    #print("texts", texts)
    if isinstance(texts, (float, int)) or texts is None:
        return ""
    if isinstance(texts, str):
        texts = texts.strip().split('\n')
 
    normalized_list = [remove_quotes(step) for step in texts]
    return normalized_list[0] if normalized_list else ""
 
def clean_test_steps(test_steps):
    # Remove numbering at the start or after new line
    cleansed_test_steps = re.sub(r'(^\d+[-|.]\s*|\n\d+[-|.]\s*)', '', test_steps, flags=re.MULTILINE)
 
    # Remove extra space spaces within the text
    cleansed_test_steps = re.sub(r'\s+', ' ', cleansed_test_steps)
    # Remove leading and trailing spaces and newlines
    cleansed_test_steps = cleansed_test_steps.strip()
    # Ensure no extra newline at the start
    cleansed_test_steps = cleansed_test_steps.lstrip('\n')
    cleansed_test_steps = cleansed_test_steps.lower()
    return cleansed_test_steps
 
def split_text(text):
    # Split the cleaned text by numbered steps
    steps = re.split(r'\n\d+[-|.]\s*', text.strip())
    steps = [re.sub(r'\s+', ' ', step).strip() for step in steps]
    if steps and steps[0].startswith('1.'):
        steps[0] = steps[0][2:].strip()
    return steps
 
def split_text_1(text):
    # Split the cleaned text by numbered steps
    steps = re.split(r'\n\d+[-|.]\s*', text.strip())
    # Remove extra spaces from each step
    steps = [re.sub(r'\s+', ' ', step).strip() for step in steps]
    # Ensure the first step starts correctly after the split
    if steps and steps[0].startswith('1.'):
        steps[0] = steps[0][2:].strip()
    # Remove duplicates while maintaining order
    list_unique = list(dict.fromkeys(steps))

    return list_unique

#----------------------------------------------------------------------#

def calculate_embeddings(text):
    if not text:
        return []
    return model.encode(text)

# Process each test case to get embeddings for title and test steps
def process_row(row):
    title = row['TestCaseDescription']
    steps = [step for step in row['TestSteps'] if isinstance(step, str) and step.strip()]
    title_embedding = model.encode([title])[0]
    steps_embedding = model.encode(steps) if steps else np.array([])  # Use an empty array if no steps
    return row['tcid'], title_embedding, steps_embedding

# Generate embeddings for all test cases using threading
def get_embedding(df1):
    desc_embeddings_dict = {} 
    step_embeddings_dict = {}

    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {executor.submit(process_row, row): row['tcid'] for _, row in df1.iterrows()}
        
        for future in as_completed(futures):
            tc_id = futures[future]
            try:
                title_embedding, steps_embedding = future.result()[1:]
                desc_embeddings_dict[tc_id] = title_embedding
                step_embeddings_dict[tc_id] = steps_embedding
            except Exception as e:
                print(f"Error processing test case {tc_id}: {e}")
    
    return desc_embeddings_dict, step_embeddings_dict


#----------------------------------------------------------------------#
def process_embedding(df1,desc_embeddings_dict,step_embeddings_dict):
    try:
        embedding_dim = model.get_sentence_embedding_dimension()
        desc_matrix = np.array([desc_embeddings_dict.get(tc_id,np.zeros(embedding_dim)) for tc_id in df1['tcid']])    
        
        step_matrix = np.array([ 
            np.mean(step_embeddings_dict.get(tc_id, [np.zeros(embedding_dim)]), axis = 0)
            if len(step_embeddings_dict.get(tc_id,[])) > 0 else np.zeros(embedding_dim)  
            for tc_id in df1['tcid']
                    
            ])
        return desc_matrix,step_matrix
    except Exception as e:
        print("Exception in process_embedding",e)

def process_similarity_batch(batch, title_sim_matrix, test_step_sim_matrix, df1, step_embeddings_dict):
    try:
        results = []
        print("----------333333333-------------------")
        for i, j in batch:
            tc_id1 = df1['tcid'].iloc[i]
            tc_id2 = df1['tcid'].iloc[j]
            
            # Retrieve title and step similarity scores
            title_score = title_sim_matrix[i, j]
            step_score = test_step_sim_matrix[i, j]
            
          
            # Calculate matching steps if both test cases have step embeddings
            if step_embeddings_dict.get(tc_id1) is not None and step_embeddings_dict.get(tc_id2) is not None:
                step_similarity_matrix = cosine_similarity(step_embeddings_dict[tc_id1], step_embeddings_dict[tc_id2])
                matching_indices = np.where(step_similarity_matrix >= 0.95)
                
                matching_steps = [
                    (df1['TestSteps'].iloc[i][step_i], df1['TestSteps'].iloc[j][step_j], step_similarity_matrix[step_i, step_j])
                    for step_i, step_j in zip(*matching_indices)
                    if step_i < len(df1['TestSteps'].iloc[i]) and step_j < len(df1['TestSteps'].iloc[j])  # Boundary check
                ]
                
                similarities = step_similarity_matrix.flatten()
                #score_list = [float(sim) for sim in similarities if sim >= 0.95 ]
                #sim_arr = np.array(similarities, dtype=np.float32)
                #score_list = [int(sim) if sim.is_integer() else float(sim) for sim in similarities if sim >= 0.95]
                score_list = [sim  for sim in similarities if sim >= 0.95]
    
            else:
                matching_steps = []
    
            results.append({
                'tc_id1': tc_id1,
                'tc_id2': tc_id2,
                'desc_score': title_score,
                'test_step_similarity': step_score,
                'matching_steps_count': len(matching_steps),
                'matching_steps' : matching_steps,
                'score_list' : score_list,
                'similarity_score': np.mean(similarities),
                'total_steps_1': df1['original_step_count'].iloc[i],
                'total_steps_2': df1['original_step_count'].iloc[j],
                'RemoveTeststepCount1': df1['step_count'].iloc[i],
                'RemoveTeststepCount2': df1['step_count'].iloc[j],
            })
    
        return results
    except Exception as e:
        print("Exception in process_similarity_batch ",e)

def calculate_similarity(df1, step_embeddings_dict,desc_embeddings_dict):
    try:
        desc_matrix,step_matrix = process_embedding(df1,desc_embeddings_dict,step_embeddings_dict)
        print("----------11111111111-------------------")
        all_pairs = list(combinations(range(len(df1)),2)) ##without redudnacy
        
        num_cores = mp.cpu_count()
        
        BATCH_SIZE = max(1,len(all_pairs) // num_cores)
        
        batches = [all_pairs[i:i + BATCH_SIZE] for i in range(0, len(all_pairs), BATCH_SIZE)]
    
        # Calculate similarity matrices for titles and test steps
        title_sim_matrix = cosine_similarity(desc_matrix) if len(desc_matrix) > 1 else np.zeros((len(df1), len(df1)))
        test_step_sim_matrix = cosine_similarity(step_matrix) if len(step_matrix) > 1 else np.zeros((len(df1), len(df1)))
        
        print("----------2222222222-------------------")
        # Process each batch in parallel
        results = []
        with ThreadPoolExecutor(max_workers=num_cores) as executor:
            future_to_batch = {executor.submit(process_similarity_batch, batch, title_sim_matrix, test_step_sim_matrix, df1, step_embeddings_dict ): batch for batch in batches}
            
            for future in future_to_batch:
                results.extend(future.result())
    
        return results
    except Exception as e:
        print("Exception in calculate_similarity ",e)


##---------------------------------------------------------------

## Get Count og Numbers which is greater than 0.99899 from score_list 
def count_greater_than_threshold(scores,threshold=0.99999):
    if len(scores) > 0 :
        scores = np.array(scores,dtype=float)
        return np.sum(scores > threshold)
    else:
        return 0

def count_greater_than_threshold_1(scores,threshold = 0.99999):
	
    if scores > threshold:
        return 1
    else:
        return scores

##---------------------------------------------------------------
## Combine score of description and test steps 
def combined_score(row):
	combined_score = [(TEST_STEP_WEIGHT * row['similarity_score']) +(TEST_CASE_WEIGHT * row['desc_score'])]
	return combined_score[0]

##---------------------------------------------------------------
# Added New Category Plus Change the logic
def calculate_duplicate_matched(row):
    highlysimilar_ratio = int(row['HighestValue'] * 0.8)
    moderate_ratio = int(row['HighestValue'] * 0.7)
    
    output_str = ''
    if ((row['title_score'] == 100) and ((row['test_step_score'] == 100) or ((row['count_of_onces'] == row['RemoveTeststepCount1']) and (row['count_of_onces'] == row['RemoveTeststepCount2'])))):
        output_str = "Duplicate"
    elif (((row['title_score'] >= 99 and row['title_score'] <= 100 ) and (row['count_of_onces'] >= row['HighestValue']-1) and row['count_of_onces'] > 0) or ((row['count_of_onces'] == row['RemoveTeststepCount1']) and (row['count_of_onces'] == row['RemoveTeststepCount2']))):
        output_str = "almostDuplicate"
    elif (((row['title_score'] >= 90 and row['title_score'] <= 100) and ((row['test_step_score'] >= 90 and row['test_step_score'] < 100) and (row['count_of_onces'] >= highlysimilar_ratio))) and (row['count_of_onces'] > 0)):
        output_str = "HighlySimilar"
    elif (((moderate_ratio <= row['count_of_onces']) or ((math.ceil(row['HighestValue']/2)) <= row['count_of_onces'])) and row['count_of_onces'] > 0):
        output_str = "moderateDuplicate"
    elif (row['count_of_onces'] == 0):
        output_str = "unique"
    else:
        output_str = "almostUnique"
    
    return output_str

##---------------------------------------------------------------
def category_summary(data_frame):
    data_frame_summary = data_frame.groupby('TC_ID1').agg(
        count=('TC_ID2', 'size'),
        DuplicateIds=('TC_ID2', lambda x: '<br>'.join(map(str, x)))
    ).reset_index()
    data_frame_summary = data_frame_summary.head(5)
    return data_frame_summary

##---------------------------------------------------------------

def mask_dataframe_output(output_df):
	#Mask duplicates
	output_df = output_df.sort_values(by='TC_ID1', ascending = False)
	output_df[['TC_ID1','Title1','TestSteps1']] = output_df[['TC_ID1','Title1','TestSteps1']].mask(output_df[['TC_ID1','Title1','TestSteps1']].duplicated(),'')
	#output_df[['TC_ID1', 'Title1', 'TestSteps1']] = output_df[['TC_ID1', 'Title1', 'TestSteps1']].mask(output_df[['TC_ID1', 'Title1', 'TestSteps1']].duplicated(keep='first'),'')
    # Set the index 
	output_df.set_index(['TC_ID1','Title1','TestSteps1'],inplace=True)
	return output_df 

##---------------------------------------------------------------
##get maskdataframe group by as per test case ids 
def get_all_columns_data(filtered_df,df1):
	#in_set : tcidl, not_in_set = tc_id2
	merged_df1 = filtered_df.merge(df1,left_on='in_set',right_on='tcid',how='left')
	merged_df1 = merged_df1.rename(columns ={'Title':'Title1','Test_Steps':'TestSteps1','tcid':'TC_ID1','original_step_count':'Test_step_count_1'})
	
	final_df = merged_df1.merge(df1,left_on='not_in_set',right_on='tcid',how='left')
	final_df = final_df.rename(columns ={'Title':'Title2','Test_Steps':'TestSteps2','tcid':'TC_ID2','original_step_count':'Test_step_count_2'})
	final_df = final_df.rename(columns = {'count_of_onces':'Matching_Step_Count','matching_steps':'Matching_Steps'})
	final_dataframe = final_df[['TC_ID1','Title1','TestSteps1','TC_ID2','Title2','TestSteps2','Test_step_count_1','Test_step_count_2','Matching_Step_Count','Matching_Steps']]

	return final_dataframe 

##---------------------------------------------------------------
def get_tc_ids_mapping(id_set, df):
    tcid1_in_set = df['tc_id1'].isin(id_set)
    tcid2_in_set = df['tc_id2'].isin(id_set)

    df['in_set'] = np.nan
    df['not_in_set'] = np.nan

    both_in_set = tcid1_in_set & tcid2_in_set

    mask_tcid1_higher = df['total_steps_1'] >= df['total_steps_2']
    mask_tcid2_higher = ~mask_tcid1_higher

    df.loc[both_in_set & mask_tcid1_higher, 'in_set'] = df['tc_id1']
    df.loc[both_in_set & mask_tcid1_higher, 'not_in_set'] = df['tc_id2']
    df.loc[both_in_set & mask_tcid2_higher, 'in_set'] = df['tc_id2']
    df.loc[both_in_set & mask_tcid2_higher, 'not_in_set'] = df['tc_id1']

    df.loc[tcid1_in_set & ~tcid2_in_set, 'in_set'] = df['tc_id1']
    df.loc[tcid1_in_set & ~tcid2_in_set, 'not_in_set'] = df['tc_id2']
    df.loc[tcid2_in_set & ~tcid1_in_set, 'in_set'] = df['tc_id2']
    df.loc[tcid2_in_set & ~tcid1_in_set, 'not_in_set'] = df['tc_id1']

    df_filtered = df.dropna(subset=['in_set'])
    return df_filtered

##---------------------------------------------------------------
def get_duplicate_data(final_dataframe, df1):
    n_duplicate_df = final_dataframe.loc[final_dataframe['result'] == 'Duplicate']
    n_tc_id1 = n_duplicate_df['tc_id1'].tolist()
    n_tc_id2 = n_duplicate_df['tc_id2'].tolist()
    n_tcids = set(n_tc_id1 + n_tc_id2)
    
    duplicate_df = get_tc_ids_mapping(n_tcids, n_duplicate_df)
    duplicate_final_df = get_all_columns_data(duplicate_df, df1)
    
    id1 = set(duplicate_final_df['TC_ID1'].tolist())
    id2 = set(duplicate_final_df['TC_ID2'].tolist())
    
    duplicate_tc1_count = len(id1)
    duplicate_tc2_count = len(id2 - id1)
    
    duplicate_data_summary = category_summary(duplicate_final_df)
    duplicate_df = mask_dataframe_output(duplicate_final_df)
    
    return duplicate_df, n_tcids, duplicate_data_summary,duplicate_tc1_count,duplicate_tc2_count

##---------------------------------------------------------------
def get_almost_duplicate(final_dataframe, n_tcids, df1):
    nl_duplicate_df = final_dataframe.loc[final_dataframe['result'] == 'almostDuplicate']
    nl_tc_id1 = nl_duplicate_df['tc_id1'].tolist()
    nl_tc_id2 = nl_duplicate_df['tc_id2'].tolist()
    nl_tc_ids = set(nl_tc_id1 + nl_tc_id2)  # Remove duplicate ids
    nl_tcids = nl_tc_ids.union(n_tcids)  # Check with previous category all ids with current and previous one
    
    n_nl_union = nl_tcids.union(nl_tcids)  # For next category: all ids with current and previous one
   
    filtered_nl_df = nl_duplicate_df[(nl_duplicate_df['tc_id1'].isin(nl_tcids) | nl_duplicate_df['tc_id2'].isin(nl_tcids))]
    almost_df = get_tc_ids_mapping(nl_tcids, filtered_nl_df)
    almost_final_df = get_all_columns_data(almost_df, df1)
    
    id1 = set(almost_final_df['TC_ID1'].tolist())
    id2 = set(almost_final_df['TC_ID2'].tolist())
    
    almost_tc1_count = len(id1)
    almost_tc2_count = len((id2 - id1) - n_tcids)
    
    almost_duplicate_summary = category_summary(almost_final_df)
    
    almost_duplicate_df = mask_dataframe_output(almost_final_df)
    return almost_duplicate_df, nl_tcids, n_nl_union, almost_duplicate_summary,almost_tc1_count,almost_tc2_count

##---------------------------------------------------------------
def get_highly_similar(final_dataframe, n_nl_union, df1):
    n2_duplicate_df = final_dataframe.loc[final_dataframe['result'] == 'highlySimilar']
    n2_tc_id1 = n2_duplicate_df['tc_id1'].tolist()
    n2_tc_id2 = n2_duplicate_df['tc_id2'].tolist()
    n2_tc_ids = set(n2_tc_id1 + n2_tc_id2)

    n2_tcids = n2_tc_ids.union(n_nl_union)
    n_nl_n2_union = n_nl_union.union(n2_tcids)

    # Gets the data as per n2_tcids which is present either tc_id1 or tc_id2
    filtered_n2_df = n2_duplicate_df[(n2_duplicate_df['tc_id1'].isin(n2_tcids) | n2_duplicate_df['tc_id2'].isin(n2_tcids))]

    highly_similar_df = get_tc_ids_mapping(n2_tcids, filtered_n2_df)
    highly_similar_final_df = get_all_columns_data(highly_similar_df, df1)
    
    id1 = set(highly_similar_final_df['TC_ID1'].tolist())
    id2 = set(highly_similar_final_df['TC_ID2'].tolist())
    
    highlysimilar_tc1_count = len(id1)
    highlysimilar_tc2_count = len((id2 - id1) - n_nl_union)
    
    highly_similar_summary = category_summary(highly_similar_final_df)
    highly_similar_duplicate_df = mask_dataframe_output(highly_similar_final_df)
    return highly_similar_duplicate_df, n2_tcids, n_nl_n2_union, highly_similar_summary,highlysimilar_tc1_count,highlysimilar_tc2_count

##-----------------------------------------------------------------------
def get_moderate_duplicate(final_dataframe, n_nl_n2_union, df1):
    n3_duplicate_df = final_dataframe.loc[final_dataframe['result'] == 'moderateDuplicate']
    n3_tc_id1 = n3_duplicate_df['tc_id1'].tolist()
    n3_tc_id2 = n3_duplicate_df['tc_id2'].tolist()
    n3_tc_ids = set(n3_tc_id1 + n3_tc_id2)

    n3_tcids = n3_tc_ids.union(n_nl_n2_union)

    n_nl_n2_n3_union = n_nl_n2_union.union(n3_tc_ids)

    filtered_n3_df = n3_duplicate_df[(n3_duplicate_df['tc_id1'].isin(n3_tcids) | n3_duplicate_df['tc_id2'].isin(n3_tcids))]
    moderate_df = get_tc_ids_mapping(n3_tcids, filtered_n3_df)
    moderate_final_df = get_all_columns_data(moderate_df, df1)

    id1 = set(moderate_final_df['TC_ID1'].tolist())
    id2 = set(moderate_final_df['TC_ID2'].tolist())
   
    moderate_tc1_count = len(id1)
    moderate_tc2_count = len((id2 - id1) - n_nl_n2_union)

    moderate_duplicate_summary = category_summary(moderate_final_df)
    moderate_duplicate_df = mask_dataframe_output(moderate_final_df)
    return moderate_duplicate_df, n3_tcids, n_nl_n2_n3_union, moderate_duplicate_summary,moderate_tc1_count,moderate_tc2_count


##-----------------------------------------------------------------------
def get_almost_unique(final_dataframe, n_nl_n2_n3_union, df1):
    n4_duplicate_df = final_dataframe.loc[final_dataframe['result'] == 'almostUnique']
    n4_tc_id1 = n4_duplicate_df['tc_id1'].tolist()
    n4_tc_id2 = n4_duplicate_df['tc_id2'].tolist()
    n4_tc_ids = set(n4_tc_id1 + n4_tc_id2)

    n4_tcids = n4_tc_ids.union(n_nl_n2_n3_union)

    n_nl_n2_n3_n4_union = n_nl_n2_n3_union.union(n4_tc_ids)

    filtered_n4_df = n4_duplicate_df[(n4_duplicate_df['tc_id1'].isin(n4_tcids) | n4_duplicate_df['tc_id2'].isin(n4_tcids))]
    almostUnique_df = get_tc_ids_mapping(n4_tcids, filtered_n4_df)
    almostUnique_df_final_df = get_all_columns_data(almostUnique_df, df1)

    id1 = set(almostUnique_df_final_df['TC_ID1'].tolist())
    id2 = set(almostUnique_df_final_df['TC_ID2'].tolist())
   
    almostunique_tc1_count = len(id1)
    almostunique_tc2_count = len((id2 - id1) - n_nl_n2_n3_union)

    almost_unique_summary = category_summary(almostUnique_df_final_df)
    almost_unique_df = mask_dataframe_output(almostUnique_df_final_df)
    return almost_unique_df, n4_tcids, n_nl_n2_n3_n4_union, almost_unique_summary,almostunique_tc1_count,almostunique_tc2_count

##------------------------------------------------------------------------
def get_unique_data(final_dataframe, n_nl_n2_n3_n4_union, df1):
    unique_df = final_dataframe.loc[final_dataframe['result'] == 'unique']
    unique_tc_id1 = unique_df['tc_id1'].tolist()
    unique_tc_id2 = unique_df['tc_id2'].tolist()
    unique_tcids = set(unique_tc_id1 + unique_tc_id2)

    n5_tcids = unique_tcids.union(n_nl_n2_n3_n4_union)

    unique_df = df1.loc[df1['tcid'].isin(n5_tcids)]

    merged_unique_df = unique_df.merge(df1, left_on='tcid', right_on='tcid', how='left')

    merged_unique_df = merged_unique_df.rename(columns={'tcid': 'TC_ID', 'Title_x': 'Title', 'Test_Steps_x': 'Test_Steps', 'step_count_x': 'Test_step_count'})

    final_merged_unique_df = merged_unique_df[['TC_ID', 'Title', 'Test_Steps', 'Test_step_count']]

    unique_summary = final_merged_unique_df.head(5)

    return final_merged_unique_df, n5_tcids, unique_summary


##-----------------------------------------------------------------------
def format_excelFile(df,filename,sheet_name):
    
    if sheet_name == "Unique":
        index = False
    else:
        index = True 
        
    with pd.ExcelWriter(filename,engine='openpyxl') as writer:
        df.to_excel(writer,sheet_name=sheet_name, index=index)
        
        workbook = writer.book
        sheet = writer.sheets[sheet_name]
        
        header_fill = PatternFill(start_color = 'FFFF00', end_color='FFFF00',fill_type='solid')
        header_font = Font(bold=True)
        left_alignment = Alignment(horizontal="left")
        
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = left_alignment
        
        for row in sheet.iter_rows(min_row=sheet.max_row,min_col=1,max_col=sheet.max_column):
            for cell in row:
                cell.alignment = left_alignment
                cell.border = Border()
                cell.font = Font(bold=False)
                
        for col in sheet.columns:
            max_length = 20
            col_letter = col[0].column_letter
            sheet.column_dimensions[col_letter].width = max_length

##-----------------------------------------------------------------------
def main(data, filename):
    try:
        #data = pd.read_excel("ForgotPasswordModule.xlsx")
        #filename = "Newchange"
        df = data[['Title', 'TestSteps', 'ID']]
        print("1.................................")
        # Check for empty or missing values in ID, Title, TestSteps
        invalid_rows = df[
            (df[['ID', 'Title', 'TestSteps']].isnull().any(axis=1) | 
             (df[['Title', 'TestSteps']] == '').any(axis=1))
        ]
        print("2.................................")
        invalid_summary = invalid_rows.head(5)
        d_df = df.drop(invalid_rows.index)
        
        df1 = pd.DataFrame()
        df1['TestCaseDescription'] = d_df['Title'].apply(preprocess)
        
        df1['original_TestSteps'] = d_df['TestSteps'].apply(split_text) # original TestSteps
        df1['original_step_count'] = df1['original_TestSteps'].apply(len)
        df1['TestSteps'] = d_df['TestSteps'].apply(split_text_1) # Remove Duplicates
        df1['step_count'] = df1['TestSteps'].apply(len)
        df1['Title'] = d_df['Title']
        df1['Test_Steps'] = d_df['TestSteps']
        df1['tcid'] = d_df['ID']
        print("3.................................")
        
        # Load a pre-trained SentenceTransformer model
        # model = SentenceTransformer('all-MiniLM-L6-v2')
        

        begin_time = datetime.datetime.now()
        desc_embeddings_dict,step_embeddings_dict = get_embedding(df1)
        end_time = datetime.datetime.now()
        print("Execution time embedding: ", end_time - begin_time)
        print("4.................................")
        begin_time1 = datetime.datetime.now()
        results = calculate_similarity(df1,step_embeddings_dict,desc_embeddings_dict)
        end_time1 = datetime.datetime.now()
        print("Execution time  similarity: ", end_time1 - begin_time1)
        print("5.................................")
        # Convert results to a DataFrame for better readability
        df_results = pd.DataFrame(results)
        
        print("6.................................")
        df_results['count_of_onces'] = df_results['score_list'].apply(count_greater_than_threshold,0.999)
        df_results['combined_score'] = df_results.apply(combined_score, axis=1)
        df_results['LowestValue'] = df_results[['RemoveTeststepCount1', 'RemoveTeststepCount2']].min(axis=1)
        df_results['HighestValue'] = df_results[['RemoveTeststepCount1', 'RemoveTeststepCount2']].max(axis=1)
        print("7.................................")
        df_results['desc_score_'] = df_results['desc_score'].apply(count_greater_than_threshold_1,0.999)
        df_results['test_step_sim_score_'] = df_results['test_step_similarity'].apply(count_greater_than_threshold_1,0.999)
        
        df_results['title_score'] = (df_results['desc_score_'] * 100).apply(np.floor).astype(int)
        df_results['test_step_score'] = (df_results['test_step_similarity'] * 100).apply(np.floor).astype(int)
        print("8.................................")
        df_results['result'] = df_results.apply(calculate_duplicate_matched, axis=1)
        print("9.................................")
        duplicate_df, n_tcids, duplicate_data_summary,duplicate_tc1_count,duplicate_tc2_count = get_duplicate_data(df_results, df1)
        almost_duplicate_df, n1_tcids, n1_union, almost_duplicate_summary,almost_tc1_count,almost_tc2_count = get_almost_duplicate(df_results, n_tcids, df1)
        highlysimilar_duplicate_df, n2_tcids, n1_n2_union, highlysimilar_summary,highlysimilar_tc1_count,highlysimilar_tc2_count = get_highly_similar(df_results, n1_union, df1)
        moderate_duplicate_df, n3_tcids, n1_n2_n3_union, moderate_duplicate_summary,moderate_tc1_count,moderate_tc2_count = get_moderate_duplicate(df_results, n1_n2_union, df1)
        almost_unique_df, n4_tcids, n1_n2_n3_n4_union, almost_unique_summary,almostunique_tc1_count,almostunique_tc2_count = get_almost_unique(df_results, n1_n2_n3_union, df1)
        unique_df, n5_tcids, unique_summary = get_unique_data(df_results, n1_n2_n3_n4_union, df1)

        dashboard_dict = {
            "Category": ["Duplicate", "Almost Duplicate", "Highly Similar", "Moderate Duplicate", "Almost Unique", "Unique","Invalid"],
            "Similarity_Percentage": ["100%", "99-100%", "75-98%", "50-74%", "25-49%", "< 25%","-"],
            "Recommendation": ["Remove", "Review (Elementary)", "Review (Moderate)", "Review (Exhuastive)","Review (Exhuastive)","Retain","Invalid"],
            "Test_Cases_Count": [len(n_tcids), len(n1_tcids), len(n2_tcids), len(n3_tcids), len(n4_tcids), len(n5_tcids), len(invalid_rows)],
            "Retain_Test_Cases_Count":[duplicate_tc1_count,almost_tc1_count,highlysimilar_tc1_count,moderate_tc1_count,almostunique_tc1_count,len(n5_tcids),0],
            "Remove_Test_Cases_Count":[duplicate_tc2_count,almost_tc2_count,highlysimilar_tc2_count,moderate_tc2_count,almostunique_tc2_count,0,0],
            "Colors": ["#E23610", "#F4914C", "#FFC885", "#FFBF00", "#93C572", "#77dd77","#808080"],
            "Explode": [0.01, 0.01, 0.01, 0.01, 0.01, 0.01,0.01]
        }

        dashboard_df = pd.DataFrame(dashboard_dict)
        dashboard_filtered_df = dashboard_df[dashboard_df['Test_Cases_Count'] != 0]

        output_dir = 'output'
        if not os.path.isdir(output_dir):
            os.mkdir(output_dir)

        format_excelFile(duplicate_df, output_dir + "/" + filename + "_DuplicateTC.xlsx", "Duplicates")
        format_excelFile(almost_duplicate_df, output_dir + "/" + filename + "_Almost_DuplicatesTC.xlsx", "Almost_Duplicates")
        format_excelFile(highlysimilar_duplicate_df, output_dir + "/" + filename + "_Highly_SimilarTC.xlsx", "Highly_Similar")
        format_excelFile(moderate_duplicate_df, output_dir + "/" + filename + "_Moderate_DuplicateTC.xlsx", "Moderate_Duplicates")
        format_excelFile(almost_unique_df, output_dir + "/" + filename + "_Almost_UniqueTC.xlsx", "Almost_Uniques")
        format_excelFile(unique_df, output_dir + "/" + filename + "_UniqueTC.xlsx", "Uniques")
        format_excelFile(invalid_rows, output_dir + "/" + filename + "_InvalidTC.xlsx", "Invalid")
        format_excelFile(dashboard_filtered_df, output_dir + "/" + filename + "_Summary.xlsx", "Summary")

        duplicate_data_summary.to_excel(output_dir + "/" + filename + "_Duplicate.xlsx")
        almost_duplicate_summary.to_excel(output_dir + "/" + filename + "_Almost_Duplicates.xlsx")
        highlysimilar_summary.to_excel(output_dir + "/" + filename + "_HighlySimilar.xlsx")
        moderate_duplicate_summary.to_excel(output_dir + "/" + filename + "_ModerateSimilar.xlsx")
        almost_unique_summary.to_excel(output_dir + "/" + filename + "_AlmostUnique.xlsx")
        unique_summary.to_excel(output_dir + "/" + filename + "_Unique.xlsx")
        invalid_summary.to_excel(output_dir + "/" + filename + "_Invalid.xlsx")
      
        #return dashboard_filtered_df
        return dashboard_filtered_df, duplicate_data_summary, almost_duplicate_summary, highlysimilar_summary, moderate_duplicate_summary,almost_unique_summary,unique_summary,invalid_summary
    except Exception as e:
        print("Exception:", e)












